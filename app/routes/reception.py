from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status as http_status, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional, List
import os
import shutil
from datetime import datetime
import uuid
import io
import json
from pydantic import BaseModel, Field
from fastapi import HTTPException, status
from pydantic import ValidationError
from app.database import get_db
from app.models import (
    Customer, Device, RepairOrder, User, StatusHistory, Attachment,
    Site, SiteType, Panel, Board, BoardType, OrderStatus
)
from app.schemas import (
    CustomerCreate, CustomerResponse,
    DeviceCreate, DeviceResponse,
    RepairOrderCreate, RepairOrderResponse
)
from app.config import settings
from app.auth import (
    get_current_user, 
    get_current_active_user,
    get_current_admin_user,
    get_current_technical_user,
    get_current_reception_user
)
from app.schemas import PhysicalCondition
from fastapi.responses import HTMLResponse
# کتابخانه‌های Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi import Request

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reception", tags=["پذیرش"])
templates = Jinja2Templates(directory="frontend/templates")
# ============================================
# مدل‌های Pydantic برای APIهای جدید
# ============================================

class SiteCreate(BaseModel):
    name: str
    type: str
    address: Optional[str] = None
    location: Optional[str] = None
    description: Optional[str] = None
    building_name: Optional[str] = None
    building_manager: Optional[str] = None
    manager_phone: Optional[str] = None
    lobby_phone: Optional[str] = None
    responsible_name: Optional[str] = None
    responsible_position: Optional[str] = None
    responsible_phone: Optional[str] = None
    customer_id: int

class SiteResponse(BaseModel):
    id: int
    name: str
    type: str
    address: Optional[str]
    location: Optional[str]
    description: Optional[str]
    building_name: Optional[str]
    building_manager: Optional[str]
    manager_phone: Optional[str]
    lobby_phone: Optional[str]
    responsible_name: Optional[str]
    responsible_position: Optional[str]
    responsible_phone: Optional[str]
    customer_id: int
    created_at: datetime
    
    class Config:
        from_attributes = True

class PanelCreate(BaseModel):
    brand: str
    model: str
    serial_number: str
    part_number: str
    firmware_version: Optional[str] = None
    hardware_version: Optional[str] = None
    loops_count: int = 0
    zones_count: int = 0
    installation_year: Optional[int] = None

class PanelResponse(BaseModel):
    id: int
    brand: str
    model: str
    serial_number: str
    part_number: str
    firmware_version: Optional[str]
    hardware_version: Optional[str]
    loops_count: int
    zones_count: int
    installation_year: Optional[int]
    created_at: datetime
    
    class Config:
        from_attributes = True

class BoardCreate(BaseModel):
    board_type: str
    part_number: str
    serial_number: str
    revision: Optional[str] = None
    description: Optional[str] = None

class RepairOrderV2Create(BaseModel):
    customer_id: int
    site_id: Optional[int] = None
    panel_id: int
    sender_name: Optional[str] = None
    sender_position: Optional[str] = None
    sender_phone: Optional[str] = None
    sender_landline: Optional[str] = None
    delivery_method: Optional[str] = None
    courier_company: Optional[str] = None
    courier_tracking: Optional[str] = None
    physical_damages: List[str] = []
    physical_description: Optional[str] = None
    accessories: List[str] = []
    accessories_description: Optional[str] = None
    customer_complaint: str
    boards: List[dict] = []

# ============================================
# مدل برای تغییر وضعیت
# ============================================
class StatusUpdateRequest(BaseModel):
    status: str
    reason: Optional[str] = None

# ============================================
# توابع کمکی
# ============================================
def generate_tracking_code() -> str:
    """تولید کد رهگیری یکتا"""
    return f"SR-{datetime.now().strftime('%Y%m')}-{uuid.uuid4().hex[:6].upper()}"

def get_status_enum(status_str: str) -> OrderStatus:
    """تبدیل وضعیت فارسی به Enum"""
    status_map = {
        "ثبت شده": OrderStatus.REGISTERED,
        "در انتظار بررسی فنی": OrderStatus.WAITING_TECHNICAL,
        "در حال عیب‌یابی": OrderStatus.DIAGNOSING,
        "در انتظار تایید مشتری": OrderStatus.WAITING_APPROVAL,
        "در حال تعمیر": OrderStatus.REPAIRING,
        "در حال تست": OrderStatus.TESTING,
        "کنترل نهایی": OrderStatus.FINAL_CONTROL,
        "آماده تحویل": OrderStatus.READY_DELIVERY,
        "تحویل شده": OrderStatus.DELIVERED,
        "مختومه بدون تعمیر": OrderStatus.CLOSED_NO_REPAIR
    }
    if status_str not in status_map:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"وضعیت '{status_str}' معتبر نیست. وضعیت‌های مجاز: {', '.join(status_map.keys())}"
        )
    return status_map[status_str]

def format_order_response(order: RepairOrder) -> dict:
    """فرمت کردن پاسخ پرونده"""
    return {
        "id": order.id,
        "tracking_code": order.tracking_code,
        "status": order.status.value if hasattr(order.status, 'value') else str(order.status),
        "reception_date": order.reception_date,
        "technical_review_date": order.technical_review_date,
        "diagnosis_date": order.diagnosis_date,
        "repair_start_date": order.repair_start_date,
        "repair_complete_date": order.repair_complete_date,
        "final_delivery_date": order.final_delivery_date,
        "customer_complaint": order.customer_complaint,
        "notes": order.notes,
        "priority": order.priority,
        "created_at": order.created_at,
        "customer_name": order.customer.name if order.customer else None,
        "customer_company": order.customer.company if order.customer else None,
        "customer_phone": order.customer.phone if order.customer else None,
        "customer_address": order.customer.address if order.customer else None,
        # ✅ تغییر: استفاده از panel به جای device
        "device_brand": order.panel.brand if order.panel else None,
        "device_model": order.panel.model if order.panel else None,
        "device_part_number": order.panel.part_number if order.panel else None,
        "device_serial_number": order.panel.serial_number if order.panel else None
    }

# ============================================
# بخش 1: APIهای قدیمی (حفظ شده)
# ============================================

# --------------------------------------------
# 1. ثبت مشتری جدید
# --------------------------------------------
@router.post("/customers", response_model=CustomerResponse, status_code=http_status.HTTP_201_CREATED)
def create_customer(
    customer: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """ثبت مشتری جدید - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"ثبت مشتری جدید توسط: {current_user.username}")
    
    existing = db.query(Customer).filter(Customer.phone == customer.phone).first()
    if existing:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="شماره تماس قبلاً ثبت شده است"
        )
    
    if customer.email:
        existing = db.query(Customer).filter(Customer.email == customer.email).first()
        if existing:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="ایمیل قبلاً ثبت شده است"
            )
    
    try:
        db_customer = Customer(**customer.model_dump())
        db.add(db_customer)
        db.commit()
        db.refresh(db_customer)
        logger.info(f"✅ مشتری جدید ثبت شد: {db_customer.name} - {db_customer.phone}")
        return db_customer
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ثبت مشتری: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ثبت مشتری: {str(e)}"
        )

# --------------------------------------------
# 2. ثبت دستگاه جدید
# --------------------------------------------
@router.post("/devices", response_model=DeviceResponse, status_code=http_status.HTTP_201_CREATED)
def create_device(
    device: DeviceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """ثبت دستگاه جدید - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"ثبت دستگاه جدید توسط: {current_user.username}")
    
    existing = db.query(Device).filter(Device.serial_number == device.serial_number).first()
    if existing:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="سریال نامبر قبلاً ثبت شده است"
        )
    
    if device.customer_id:
        customer = db.query(Customer).filter(Customer.id == device.customer_id).first()
        if not customer:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="مشتری مورد نظر یافت نشد"
            )
    
    try:
        db_device = Device(**device.model_dump())
        db.add(db_device)
        db.commit()
        db.refresh(db_device)
        logger.info(f"✅ دستگاه جدید ثبت شد: {db_device.brand} {db_device.model} - SN: {db_device.serial_number}")
        return db_device
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ثبت دستگاه: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ثبت دستگاه: {str(e)}"
        )

# --------------------------------------------
# 3. آپلود عکس دستگاه
# --------------------------------------------
@router.post("/devices/{device_id}/upload-photo")
def upload_device_photo(
    device_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """آپلود عکس دستگاه - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"آپلود عکس دستگاه {device_id} توسط: {current_user.username}")
    
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(
            status_code=http_status.HTTP_404_NOT_FOUND,
            detail="دستگاه مورد نظر یافت نشد"
        )
    
    if file.size and file.size > settings.max_file_size:
        raise HTTPException(
            status_code=http_status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"حجم فایل بیش از حد مجاز ({settings.max_file_size // (1024*1024)} مگابایت) است"
        )
    
    try:
        upload_dir = os.path.join(settings.upload_dir, "photos", str(device_id))
        os.makedirs(upload_dir, exist_ok=True)
        
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(upload_dir, unique_filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        if device.photo_paths is None:
            device.photo_paths = []
        relative_path = f"/uploads/photos/{device_id}/{unique_filename}"
        device.photo_paths.append(relative_path)
        db.commit()
        
        logger.info(f"✅ عکس دستگاه {device_id} آپلود شد")
        return {"message": "عکس با موفقیت آپلود شد", "path": relative_path}
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در آپلود عکس: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در آپلود عکس: {str(e)}"
        )

# --------------------------------------------
# 4. ایجاد پرونده تعمیر جدید (نسخه قدیمی)
# --------------------------------------------
@router.post("/repair-orders", response_model=RepairOrderResponse, status_code=http_status.HTTP_201_CREATED)
def create_repair_order(
    order: RepairOrderCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """ایجاد پرونده تعمیر جدید - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"ایجاد پرونده جدید توسط: {current_user.username}")
    
    try:
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        if not customer:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="مشتری مورد نظر یافت نشد"
            )
        
        device = db.query(Device).filter(Device.id == order.device_id).first()
        if not device:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="دستگاه مورد نظر یافت نشد"
            )
        
        tracking_code = generate_tracking_code()
        
        db_order = RepairOrder(
            **order.model_dump(),
            tracking_code=tracking_code,
            status=OrderStatus.REGISTERED
        )
        db.add(db_order)
        db.commit()
        db.refresh(db_order)
        
        logger.info(f"✅ پرونده جدید ثبت شد: {db_order.tracking_code}")
        return db_order
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ایجاد پرونده: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ایجاد پرونده: {str(e)}"
        )

# --------------------------------------------
# 5. دریافت لیست پرونده‌ها
# --------------------------------------------
@router.get("/repair-orders")
def get_repair_orders(
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """دریافت لیست پرونده‌ها - عمومی (بدون نیاز به احراز هویت)"""
    try:
        query = db.query(RepairOrder).options(
            joinedload(RepairOrder.customer),
            joinedload(RepairOrder.panel),  # ✅ تغییر از device به panel
            joinedload(RepairOrder.site)    # ✅ اضافه کردن site
        )
        
        if status:
            try:
                status_enum = get_status_enum(status)
                query = query.filter(RepairOrder.status == status_enum)
            except HTTPException:
                pass
        
        orders = query.order_by(RepairOrder.created_at.desc()).offset(skip).limit(limit).all()
        result = [format_order_response(order) for order in orders]
        return result
    
    except Exception as e:
        logger.error(f"خطا در دریافت لیست پرونده‌ها: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 6. جستجوی پیشرفته پرونده‌ها
# --------------------------------------------
# --------------------------------------------
# 6. جستجوی پیشرفته پرونده‌ها
# --------------------------------------------
@router.get("/repair-orders/search")
def search_repair_orders(
    q: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """جستجوی پیشرفته پرونده‌ها - عمومی (بدون نیاز به احراز هویت)"""
    try:
        # ✅ تغییر: استفاده از Panel به جای Device
        query = db.query(RepairOrder).join(Customer, RepairOrder.customer_id == Customer.id)\
                                     .join(Panel, RepairOrder.panel_id == Panel.id)
        
        if q:
            q_filter = f"%{q}%"
            query = query.filter(
                (Customer.name.ilike(q_filter)) |
                (Customer.phone.ilike(q_filter)) |
                (RepairOrder.tracking_code.ilike(q_filter)) |
                (Panel.serial_number.ilike(q_filter)) |
                (Panel.part_number.ilike(q_filter))
            )
        
        if status:
            try:
                status_enum = get_status_enum(status)
                query = query.filter(RepairOrder.status == status_enum)
            except HTTPException:
                pass
        
        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(RepairOrder.created_at >= date_from_parsed)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, "%Y-%m-%d")
                query = query.filter(RepairOrder.created_at <= date_to_parsed)
            except ValueError:
                pass
        
        results = query.order_by(RepairOrder.created_at.desc()).all()
        return [format_order_response(order) for order in results]
    
    except Exception as e:
        logger.error(f"خطا در جستجوی پرونده‌ها: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در جستجوی اطلاعات"
        )
# --------------------------------------------
# 7. دریافت جزئیات یک پرونده
# --------------------------------------------
@router.get("/repair-orders/{order_id}")
def get_repair_order(
    order_id: int,
    db: Session = Depends(get_db)
):
    """دریافت جزئیات یک پرونده - عمومی (بدون نیاز به احراز هویت)"""
    try:
        order = db.query(RepairOrder).options(
            joinedload(RepairOrder.customer),
            joinedload(RepairOrder.panel),  # ✅ تغییر از device به panel
            joinedload(RepairOrder.site)    # ✅ اضافه کردن site
        ).filter(RepairOrder.id == order_id).first()
        
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده مورد نظر یافت نشد"
            )
        
        return format_order_response(order)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت جزئیات پرونده {order_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 8. جستجوی مشتریان (ساده)
# --------------------------------------------
@router.get("/customers/search", response_model=List[CustomerResponse])
def search_customers_simple(
    q: Optional[str] = None,
    phone: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """جستجوی مشتریان - عمومی (بدون نیاز به احراز هویت)"""
    try:
        query = db.query(Customer)
        
        if q:
            query = query.filter(
                (Customer.name.ilike(f"%{q}%")) |
                (Customer.company.ilike(f"%{q}%"))
            )
        
        if phone:
            query = query.filter(Customer.phone.ilike(f"%{phone}%"))
        
        return query.limit(20).all()
    
    except Exception as e:
        logger.error(f"خطا در جستجوی مشتریان: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در جستجوی اطلاعات"
        )

# --------------------------------------------
# 9. جستجوی دستگاه‌ها
# --------------------------------------------
@router.get("/devices/search", response_model=List[DeviceResponse])
def search_devices(
    q: Optional[str] = None,
    serial: Optional[str] = None,
    part_number: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """جستجوی دستگاه‌ها - عمومی (بدون نیاز به احراز هویت)"""
    try:
        query = db.query(Device)
        
        if q:
            query = query.filter(
                (Device.brand.ilike(f"%{q}%")) | 
                (Device.model.ilike(f"%{q}%"))
            )
        
        if serial:
            query = query.filter(Device.serial_number.ilike(f"%{serial}%"))
        
        if part_number:
            query = query.filter(Device.part_number.ilike(f"%{part_number}%"))
        
        return query.limit(20).all()
    
    except Exception as e:
        logger.error(f"خطا در جستجوی دستگاه‌ها: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در جستجوی اطلاعات"
        )

# --------------------------------------------
# 10. تغییر وضعیت پرونده
# --------------------------------------------
@router.put("/repair-orders/{order_id}/status")
def update_order_status(
    order_id: int,
    request: StatusUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_technical_user)
):
    """تغییر وضعیت پرونده - فقط نقش فنی و ادمین"""
    logger.info(f"تغییر وضعیت پرونده {order_id} توسط: {current_user.username}")
    
    try:
        order = db.query(RepairOrder).filter(RepairOrder.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده مورد نظر یافت نشد"
            )
        
        new_status = get_status_enum(request.status)
        old_status = order.status
        
        if old_status == new_status:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail=f"پرونده در حال حاضر در وضعیت '{request.status}' است"
            )
        
        order.status = new_status
        
        if new_status == OrderStatus.WAITING_TECHNICAL:
            order.technical_review_date = datetime.now()
        elif new_status == OrderStatus.DIAGNOSING:
            order.diagnosis_date = datetime.now()
        elif new_status == OrderStatus.REPAIRING:
            order.repair_start_date = datetime.now()
        elif new_status == OrderStatus.TESTING:
            order.repair_complete_date = datetime.now()
        elif new_status == OrderStatus.READY_DELIVERY:
            pass
        elif new_status == OrderStatus.DELIVERED:
            order.final_delivery_date = datetime.now()
        
        history = StatusHistory(
            repair_order_id=order.id,
            old_status=old_status,
            new_status=new_status,
            reason=request.reason or f"تغییر وضعیت از {old_status.value} به {new_status.value}",
            changed_by=current_user.id,
            operator_name=current_user.full_name
        )
        db.add(history)
        db.commit()
        db.refresh(order)
        
        logger.info(f"✅ وضعیت پرونده {order.tracking_code} از {old_status.value} به {new_status.value} تغییر کرد")
        
        return {
            "message": "وضعیت با موفقیت بروزرسانی شد",
            "order_id": order.id,
            "tracking_code": order.tracking_code,
            "old_status": old_status.value,
            "new_status": new_status.value,
            "reason": request.reason,
            "operator": current_user.full_name
        }
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در تغییر وضعیت پرونده {order_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در تغییر وضعیت: {str(e)}"
        )

# --------------------------------------------
# 11. دریافت تاریخچه تغییرات
# --------------------------------------------
@router.get("/repair-orders/{order_id}/history")
def get_order_history(
    order_id: int,
    db: Session = Depends(get_db)
):
    """دریافت تاریخچه تغییرات پرونده - عمومی (بدون نیاز به احراز هویت)"""
    try:
        order = db.query(RepairOrder).filter(RepairOrder.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده مورد نظر یافت نشد"
            )
        
        history = db.query(StatusHistory).filter(
            StatusHistory.repair_order_id == order_id
        ).order_by(StatusHistory.changed_at.desc()).all()
        
        return history
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت تاریخچه پرونده {order_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 12. دریافت آمار پرونده‌ها
# --------------------------------------------
@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    """دریافت آمار پرونده‌ها - عمومی (بدون نیاز به احراز هویت)"""
    try:
        total = db.query(RepairOrder).count()
        
        repairing = db.query(RepairOrder).filter(
            RepairOrder.status.in_([
                OrderStatus.WAITING_TECHNICAL,
                OrderStatus.DIAGNOSING,
                OrderStatus.REPAIRING,
                OrderStatus.TESTING,
                OrderStatus.FINAL_CONTROL
            ])
        ).count()
        
        waiting_approval = db.query(RepairOrder).filter(
            RepairOrder.status == OrderStatus.WAITING_APPROVAL
        ).count()
        
        ready_delivery = db.query(RepairOrder).filter(
            RepairOrder.status == OrderStatus.READY_DELIVERY
        ).count()
        
        delivered = db.query(RepairOrder).filter(
            RepairOrder.status == OrderStatus.DELIVERED
        ).count()
        
        closed = db.query(RepairOrder).filter(
            RepairOrder.status == OrderStatus.CLOSED_NO_REPAIR
        ).count()
        
        return {
            "total": total,
            "repairing": repairing,
            "waiting_approval": waiting_approval,
            "ready_delivery": ready_delivery,
            "delivered": delivered,
            "closed": closed,
            "active": total - delivered - closed
        }
    
    except Exception as e:
        logger.error(f"خطا در دریافت آمار: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت آمار"
        )

# --------------------------------------------
# 13. خروجی Excel
# --------------------------------------------
@router.get("/repair-orders/export/excel")
def export_repair_orders_excel(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """خروجی Excel از پرونده‌ها - عمومی (بدون نیاز به احراز هویت)"""
    try:
        query = db.query(RepairOrder).options(
            joinedload(RepairOrder.customer),
            joinedload(RepairOrder.device)
        )
        
        if status:
            try:
                status_enum = get_status_enum(status)
                query = query.filter(RepairOrder.status == status_enum)
            except HTTPException:
                pass
        
        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(RepairOrder.created_at >= date_from_parsed)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, "%Y-%m-%d")
                query = query.filter(RepairOrder.created_at <= date_to_parsed)
            except ValueError:
                pass
        
        orders = query.order_by(RepairOrder.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "پرونده‌های تعمیرات"
        
        header_font = Font(name='B Nazanin', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            "شناسه", "کد رهگیری", "وضعیت", "تاریخ پذیرش",
            "نام مشتری", "شرکت", "شماره تماس", "آدرس",
            "برند", "مدل", "پارت نامبر", "سریال نامبر",
            "شرح مشکل", "اولویت"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=order.id)
            ws.cell(row=row_idx, column=2, value=order.tracking_code)
            ws.cell(row=row_idx, column=3, value=order.status.value if hasattr(order.status, 'value') else str(order.status))
            ws.cell(row=row_idx, column=4, value=order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "")
            ws.cell(row=row_idx, column=5, value=order.customer.name if order.customer else "")
            ws.cell(row=row_idx, column=6, value=order.customer.company if order.customer else "")
            ws.cell(row=row_idx, column=7, value=order.customer.phone if order.customer else "")
            ws.cell(row=row_idx, column=8, value=order.customer.address if order.customer else "")
            ws.cell(row=row_idx, column=9, value=order.device.brand if order.device else "")
            ws.cell(row=row_idx, column=10, value=order.device.model if order.device else "")
            ws.cell(row=row_idx, column=11, value=order.device.part_number if order.device else "")
            ws.cell(row=row_idx, column=12, value=order.device.serial_number if order.device else "")
            ws.cell(row=row_idx, column=13, value=order.customer_complaint or "")
            ws.cell(row=row_idx, column=14, value=order.priority or 0)
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        column_widths = [10, 20, 20, 25, 20, 25, 20, 35, 20, 20, 25, 25, 40, 10]
        for col, width in enumerate(column_widths, 1):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + (col-1)//26) + chr(64 + (col-1)%26 + 1)
            ws.column_dimensions[col_letter].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"گزارش_پرونده‌ها_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        encoded_filename = filename.encode('utf-8').decode('latin-1', errors='ignore')
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={encoded_filename}"}
        )
    
    except Exception as e:
        logger.error(f"خطا در خروجی Excel: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در ایجاد فایل Excel"
        )

# --------------------------------------------
# 14. آپلود فایل برای پرونده
# --------------------------------------------
@router.post("/repair-orders/{order_id}/upload")
async def upload_file(
    order_id: int,
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """آپلود فایل ضمیمه برای پرونده - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"آپلود فایل برای پرونده {order_id} توسط: {current_user.username}")
    
    try:
        order = db.query(RepairOrder).filter(RepairOrder.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده مورد نظر یافت نشد"
            )
        
        if file.size and file.size > settings.max_file_size:
            raise HTTPException(
                status_code=http_status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"حجم فایل بیش از حد مجاز ({settings.max_file_size // (1024*1024)} مگابایت) است"
            )
        
        upload_dir = os.path.join(settings.upload_dir, "attachments", str(order_id))
        os.makedirs(upload_dir, exist_ok=True)
        
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(upload_dir, unique_filename)
        
        file_size = 0
        with open(file_path, "wb") as buffer:
            while chunk := await file.read(8192):
                file_size += len(chunk)
                buffer.write(chunk)
        
        file_type = "document"
        if file.content_type and file.content_type.startswith("image/"):
            file_type = "photo"
        elif file.content_type and file.content_type.startswith("video/"):
            file_type = "video"
        elif file.content_type and file.content_type == "application/pdf":
            file_type = "pdf"
        
        attachment = Attachment(
            file_name=file.filename,
            file_path=f"/uploads/attachments/{order_id}/{unique_filename}",
            file_type=file_type,
            file_size=file_size,
            mime_type=file.content_type,
            description=description,
            repair_order_id=order_id,
            uploaded_by=current_user.id
        )
        db.add(attachment)
        db.commit()
        db.refresh(attachment)
        
        logger.info(f"✅ فایل {file.filename} برای پرونده {order_id} آپلود شد")
        
        return {
            "message": "فایل با موفقیت آپلود شد",
            "attachment_id": attachment.id,
            "file_name": attachment.file_name,
            "file_path": attachment.file_path,
            "file_type": attachment.file_type,
            "file_size": attachment.file_size
        }
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        if os.path.exists(file_path):
            os.remove(file_path)
        logger.error(f"خطا در آپلود فایل: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در آپلود فایل: {str(e)}"
        )

# --------------------------------------------
# 15. دریافت لیست فایل‌های یک پرونده
# --------------------------------------------
@router.get("/repair-orders/{order_id}/attachments")
def get_attachments(
    order_id: int,
    db: Session = Depends(get_db)
):
    """دریافت لیست فایل‌های ضمیمه - عمومی (بدون نیاز به احراز هویت)"""
    try:
        order = db.query(RepairOrder).filter(RepairOrder.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده مورد نظر یافت نشد"
            )
        
        attachments = db.query(Attachment).filter(
            Attachment.repair_order_id == order_id
        ).order_by(Attachment.uploaded_at.desc()).all()
        
        return attachments
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت فایل‌های پرونده {order_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 16. حذف فایل ضمیمه
# --------------------------------------------
@router.delete("/attachments/{attachment_id}")
def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """حذف فایل ضمیمه - نیاز به نقش پذیرش یا ادمین"""
    logger.info(f"حذف فایل {attachment_id} توسط: {current_user.username}")
    
    try:
        attachment = db.query(Attachment).filter(Attachment.id == attachment_id).first()
        if not attachment:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="فایل مورد نظر یافت نشد"
            )
        
        file_path = os.path.join(settings.upload_dir, attachment.file_path.lstrip("/uploads/"))
        if os.path.exists(file_path):
            os.remove(file_path)
        
        db.delete(attachment)
        db.commit()
        
        logger.info(f"✅ فایل {attachment_id} حذف شد")
        return {"message": "فایل با موفقیت حذف شد"}
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در حذف فایل {attachment_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در حذف فایل: {str(e)}"
        )

# --------------------------------------------
# 17. API تست
# --------------------------------------------
@router.get("/ping")
def ping(db: Session = Depends(get_db)):
    """API تست برای بررسی سلامت"""
    try:
        from sqlalchemy import text
        db.execute(text("SELECT 1"))
        return {
            "message": "pong",
            "status": "API is working!",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "message": "pong",
            "status": "Database error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

# ============================================
# بخش 2: APIهای جدید (برای فرم نسخه ۲)
# ============================================

# ============================================
# جستجوی پیشرفته مشتریان (برای فرم جدید)
# ============================================
@router.get("/customers/search-v2")
def search_customers_advanced(
    q: Optional[str] = None,
    phone: Optional[str] = None,
    name: Optional[str] = None,
    company: Optional[str] = None,
    id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    جستجوی پیشرفته مشتریان (برای فرم جدید)
    """
    try:
        query = db.query(Customer)
        
        # اگر id داده شده، مشتری رو برگردون
        if id:
            customer = query.filter(Customer.id == id).first()
            if customer:
                return customer
            return None
        
        # ✅ جستجو با شماره موبایل
        if phone:
            phone = phone.strip()
            results = query.filter(Customer.phone.ilike(f"%{phone}%")).limit(20).all()
            return results
        
        # ✅ جستجو با نام شخص
        if name:
            name = name.strip()
            name_filter = f"%{name}%"
            results = query.filter(
                (Customer.name.ilike(name_filter)) |
                (Customer.last_name.ilike(name_filter))
            ).limit(20).all()
            return results
        
        # ✅ جستجو با نام شرکت
        if company:
            company = company.strip()
            company_filter = f"%{company}%"
            results = query.filter(Customer.company.ilike(company_filter)).limit(20).all()
            return results
        
        # ✅ جستجو عمومی (در همه فیلدها)
        if q:
            q = q.strip()
            q_filter = f"%{q}%"
            results = query.filter(
                (Customer.name.ilike(q_filter)) |
                (Customer.company.ilike(q_filter)) |
                (Customer.phone.ilike(q_filter)) |
                (Customer.email.ilike(q_filter))
            ).limit(20).all()
            return results
        
        return []
        
    except Exception as e:
        print("ERROR in search_customers_advanced:", e)
        raise HTTPException(status_code=500, detail=str(e))
# --------------------------------------------
# 19. دریافت سایت‌های مشتری
# --------------------------------------------
@router.get("/customers/{customer_id}/sites", response_model=List[SiteResponse])
def get_customer_sites(
    customer_id: int,
    db: Session = Depends(get_db)
):
    """
    دریافت لیست محل‌های نصب یک مشتری
    """
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="مشتری یافت نشد"
            )
        
        sites = db.query(Site).filter(Site.customer_id == customer_id).all()
        return sites
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت سایت‌های مشتری {customer_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 20. ثبت محل نصب جدید
# --------------------------------------------
@router.post("/sites", response_model=SiteResponse, status_code=http_status.HTTP_201_CREATED)
def create_site(
    site_data: SiteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """
    ثبت محل نصب جدید
    """
    try:
        customer = db.query(Customer).filter(Customer.id == site_data.customer_id).first()
        if not customer:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="مشتری یافت نشد"
            )
        
        new_site = Site(**site_data.model_dump())
        db.add(new_site)
        db.commit()
        db.refresh(new_site)
        
        logger.info(f"✅ محل نصب جدید ثبت شد: {new_site.name} - مشتری: {customer.name}")
        return new_site
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ثبت محل نصب: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ثبت محل نصب: {str(e)}"
        )

# --------------------------------------------
# 21. ثبت پنل جدید
# --------------------------------------------
@router.post("/panels", response_model=PanelResponse, status_code=http_status.HTTP_201_CREATED)
def create_panel(
    panel_data: PanelCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """
    ثبت پنل جدید
    """
    try:
        existing = db.query(Panel).filter(Panel.serial_number == panel_data.serial_number).first()
        if existing:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="سریال نامبر قبلاً ثبت شده است"
            )
        
        new_panel = Panel(**panel_data.model_dump())
        db.add(new_panel)
        db.commit()
        db.refresh(new_panel)
        
        logger.info(f"✅ پنل جدید ثبت شد: {new_panel.brand} {new_panel.model} - SN: {new_panel.serial_number}")
        return new_panel
    
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ثبت پنل: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ثبت پنل: {str(e)}"
        )

# --------------------------------------------
# 22. جستجوی پنل‌ها
# --------------------------------------------
@router.get("/panels/search")
def search_panels(
    q: Optional[str] = None,
    type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    جستجوی پنل‌ها
    """
    try:
        query = db.query(Panel)
        
        if q:
            if type == "serial":
                query = query.filter(Panel.serial_number.ilike(f"%{q}%"))
            elif type == "part":
                query = query.filter(Panel.part_number.ilike(f"%{q}%"))
            else:
                query = query.filter(
                    (Panel.brand.ilike(f"%{q}%")) |
                    (Panel.model.ilike(f"%{q}%")) |
                    (Panel.serial_number.ilike(f"%{q}%")) |
                    (Panel.part_number.ilike(f"%{q}%"))
                )
        
        return query.limit(20).all()
    
    except Exception as e:
        logger.error(f"خطا در جستجوی پنل‌ها: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در جستجوی اطلاعات"
        )

# --------------------------------------------
# 23. دریافت پنل با ID
# --------------------------------------------
@router.get("/panels/{panel_id}", response_model=PanelResponse)
def get_panel(
    panel_id: int,
    db: Session = Depends(get_db)
):
    """
    دریافت اطلاعات یک پنل
    """
    try:
        panel = db.query(Panel).filter(Panel.id == panel_id).first()
        if not panel:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پنل یافت نشد"
            )
        return panel
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت پنل {panel_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات"
        )

# --------------------------------------------
# 24. ثبت پرونده نسخه ۲ (با تمام اطلاعات جدید)
# --------------------------------------------
@router.post("/repair-orders-v2", status_code=http_status.HTTP_201_CREATED)
async def create_repair_order_v2(
    data: str = Form(...),
    photos: List[UploadFile] = File([]),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_reception_user)
):
    """
    ثبت پرونده جدید با تمام اطلاعات (نسخه ۲) - نسخه دیباگ
    """
    import traceback
    import json
    
    try:
        # ============================================
        # مرحله 1: دریافت و parse کردن داده
        # ============================================
        logger.info("=" * 60)
        logger.info("🔍 شروع فرآیند ثبت پرونده نسخه ۲")
        logger.info(f"👤 کاربر: {current_user.username} (ID: {current_user.id})")
        
        order_data = json.loads(data)
        logger.info(f"✅ داده با موفقیت parse شد: {list(order_data.keys())}")
        
        # ============================================
        # مرحله 2: اعتبارسنجی
        # ============================================
        logger.info("🔍 شروع اعتبارسنجی...")
        
        # 2.1 بررسی مشتری
        customer_id = order_data.get('customer_id')
        logger.info(f"  - customer_id: {customer_id}")
        if not customer_id:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="لطفاً مشتری را انتخاب کنید"
            )
        
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="مشتری انتخاب شده در سیستم وجود ندارد"
            )
        logger.info(f"  ✅ مشتری پیدا شد: {customer.name} (ID: {customer.id})")
        
        # 2.2 بررسی پنل
        panel_id = order_data.get('panel_id')
        logger.info(f"  - panel_id: {panel_id}")
        if not panel_id:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="لطفاً پنل را انتخاب کنید"
            )
        
        panel = db.query(Panel).filter(Panel.id == panel_id).first()
        if not panel:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پنل انتخاب شده در سیستم وجود ندارد"
            )
        logger.info(f"  ✅ پنل پیدا شد: {panel.brand} {panel.model} (ID: {panel.id})")
        
        # 2.3 بررسی سایت (اختیاری)
        site_id = order_data.get('site_id')
        logger.info(f"  - site_id: {site_id}")
        if site_id:
            site = db.query(Site).filter(Site.id == site_id).first()
            if not site:
                raise HTTPException(
                    status_code=http_status.HTTP_404_NOT_FOUND,
                    detail="محل نصب انتخاب شده در سیستم وجود ندارد"
                )
            logger.info(f"  ✅ سایت پیدا شد: {site.name} (ID: {site.id})")
        
        # 2.4 بررسی شرح مشتری
        customer_complaint = order_data.get('customer_complaint', '').strip()
        logger.info(f"  - customer_complaint: {customer_complaint[:50]}...")
        if not customer_complaint or len(customer_complaint) < 3:
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="لطفاً شرح مشکل را وارد کنید (حداقل ۳ کاراکتر)"
            )
        logger.info("  ✅ شرح مشتری معتبر است")
        
        # ============================================
        # مرحله 3: ساخت Tracking Code و پرونده
        # ============================================
        logger.info("🔍 ساخت پرونده جدید...")
        
        tracking_code = f"SR-{datetime.now().strftime('%Y%m')}-{uuid.uuid4().hex[:6].upper()}"
        logger.info(f"  - tracking_code: {tracking_code}")
        
        # ============================================
        # مرحله 4: آماده‌سازی داده برای INSERT
        # ============================================
        repair_order_data = {
            'tracking_code': tracking_code,
            'qr_code': None,
            'status': OrderStatus.REGISTERED,
            'operator_name': current_user.full_name,
            'customer_id': customer_id,
            'site_id': site_id,
            'panel_id': panel_id,
            # ✅ device_id را حذف کردیم
            'sender_name': order_data.get('sender_name'),
            'sender_position': order_data.get('sender_position'),
            'sender_phone': order_data.get('sender_phone'),
            'sender_landline': order_data.get('sender_landline'),
            'delivery_method': order_data.get('delivery_method'),
            'courier_company': order_data.get('courier_company'),
            'courier_tracking': order_data.get('courier_tracking'),
            'physical_damages': order_data.get('physical_damages', []),
            'physical_description': order_data.get('physical_description'),
            'accessories': order_data.get('accessories', []),
            'accessories_description': order_data.get('accessories_description'),
            'customer_complaint': customer_complaint,
            'notes': None,
            'priority': 0,
            'technical_review_date': None,
            'diagnosis_date': None,
            'repair_start_date': None,
            'repair_complete_date': None,
            'final_delivery_date': None,
            'updated_at': None
        }
        
        logger.info(f"📋 داده‌های آماده برای INSERT: {repair_order_data}")
        
        # ============================================
        # مرحله 5: ایجاد پرونده در دیتابیس
        # ============================================
        logger.info("🔍 اجرای INSERT در دیتابیس...")
        
        try:
            repair_order = RepairOrder(**repair_order_data)
            db.add(repair_order)
            db.flush()
            logger.info(f"  ✅ پرونده با ID {repair_order.id} ایجاد شد")
        except Exception as insert_error:
            logger.error(f"  ❌ خطا در INSERT: {insert_error}")
            logger.error(f"  📄 جزییات: {traceback.format_exc()}")
            raise HTTPException(
                status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"خطا در ایجاد پرونده: {str(insert_error)}"
            )
        
        # ============================================
        # مرحله 6: ثبت بردها
        # ============================================
        boards_data = order_data.get('boards', [])
        logger.info(f"🔍 ثبت {len(boards_data)} برد...")
        
        for idx, board_data in enumerate(boards_data):
            try:
                board = Board(
                    board_type=board_data.get('type'),
                    part_number=board_data.get('part_number'),
                    serial_number=board_data.get('serial_number'),
                    revision=board_data.get('revision'),
                    repair_order_id=repair_order.id
                )
                db.add(board)
                logger.info(f"  ✅ برد {idx+1} اضافه شد: {board.part_number}")
            except Exception as board_error:
                logger.error(f"  ❌ خطا در ثبت برد {idx+1}: {board_error}")
                raise HTTPException(
                    status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"خطا در ثبت برد {idx+1}: {str(board_error)}"
                )
        
        # ============================================
        # مرحله 7: آپلود عکس‌ها
        # ============================================
        uploaded_photos = []
        logger.info(f"🔍 آپلود {len(photos)} عکس...")
        
        if photos:
            upload_dir = os.path.join(settings.upload_dir, "photos", str(repair_order.id))
            os.makedirs(upload_dir, exist_ok=True)
            logger.info(f"  📁 پوشه آپلود: {upload_dir}")
            
            for idx, photo in enumerate(photos):
                try:
                    logger.info(f"  📸 پردازش عکس {idx+1}: {photo.filename}")
                    file_extension = os.path.splitext(photo.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    file_path = os.path.join(upload_dir, unique_filename)
                    
                    with open(file_path, "wb") as buffer:
                        shutil.copyfileobj(photo.file, buffer)
                    
                    relative_path = f"/uploads/photos/{repair_order.id}/{unique_filename}"
                    uploaded_photos.append(relative_path)
                    
                    attachment = Attachment(
                        file_name=photo.filename,
                        file_path=relative_path,
                        file_type="photo",
                        file_size=os.path.getsize(file_path),
                        mime_type=photo.content_type,
                        is_physical_damage=True,
                        repair_order_id=repair_order.id,
                        uploaded_by=current_user.id
                    )
                    db.add(attachment)
                    logger.info(f"  ✅ عکس {idx+1} آپلود شد: {relative_path}")
                except Exception as photo_error:
                    logger.error(f"  ❌ خطا در آپلود عکس {idx+1}: {photo_error}")
                    # ادامه می‌دهیم حتی اگر یک عکس مشکل داشته باشد
        
        # ============================================
        # مرحله 8: ثبت تاریخچه
        # ============================================
        logger.info("🔍 ثبت تاریخچه وضعیت...")
        try:
            history = StatusHistory(
                repair_order_id=repair_order.id,
                old_status=None,
                new_status=OrderStatus.REGISTERED,
                reason="ثبت پرونده جدید",
                operator_name=current_user.full_name,
                changed_by=current_user.id
            )
            db.add(history)
            logger.info("  ✅ تاریخچه ثبت شد")
        except Exception as history_error:
            logger.error(f"  ❌ خطا در ثبت تاریخچه: {history_error}")
            raise HTTPException(
                status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"خطا در ثبت تاریخچه: {str(history_error)}"
            )
        
        # ============================================
        # مرحله 9: Commit نهایی
        # ============================================
        logger.info("🔍 اجرای COMMIT نهایی...")
        try:
            db.commit()
            db.refresh(repair_order)
            logger.info(f"  ✅ COMMIT موفق! پرونده {repair_order.tracking_code} ثبت شد")
        except Exception as commit_error:
            logger.error(f"  ❌ خطا در COMMIT: {commit_error}")
            logger.error(f"  📄 جزییات: {traceback.format_exc()}")
            db.rollback()
            raise HTTPException(
                status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"خطا در ذخیره‌سازی نهایی: {str(commit_error)}"
            )
        
        # ============================================
        # مرحله 10: بازگشت پاسخ موفق
        # ============================================
        logger.info("=" * 60)
        logger.info(f"✅ پرونده نسخه ۲ با موفقیت ثبت شد: {repair_order.tracking_code}")
        logger.info("=" * 60)
        
        return {
            "id": repair_order.id,
            "tracking_code": repair_order.tracking_code,
            "status": repair_order.status.value,
            "message": "پرونده با موفقیت ثبت شد",
            "photos_count": len(uploaded_photos),
            "boards_count": len(boards_data)
        }
    
    except HTTPException:
        db.rollback()
        logger.error(f"❌ خطای HTTP: {traceback.format_exc()}")
        raise
    
    except json.JSONDecodeError as e:
        db.rollback()
        logger.error(f"❌ خطا در pars کردن JSON: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"اطلاعات ارسالی نامعتبر است: {str(e)}"
        )
    
    except Exception as e:
        db.rollback()
        logger.error("=" * 60)
        logger.error("❌❌❌ خطای پیش‌بینی‌نشده در ثبت پرونده ❌❌❌")
        logger.error(f"نوع خطا: {type(e).__name__}")
        logger.error(f"پیام خطا: {str(e)}")
        logger.error("جزییات کامل:")
        logger.error(traceback.format_exc())
        logger.error("=" * 60)
        
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در ثبت پرونده: {type(e).__name__} - {str(e)}"
        )
        
        # ============================================
        # ثبت بردها
        # ============================================
        
        boards_data = order_data.get('boards', [])
        for board_data in boards_data:
            board = Board(
                board_type=board_data.get('type'),
                part_number=board_data.get('part_number'),
                serial_number=board_data.get('serial_number'),
                revision=board_data.get('revision'),
                repair_order_id=repair_order.id
            )
            db.add(board)
        
        # ============================================
        # آپلود عکس‌ها
        # ============================================
        
        uploaded_photos = []
        if photos:
            upload_dir = os.path.join(settings.upload_dir, "photos", str(repair_order.id))
            os.makedirs(upload_dir, exist_ok=True)
            
            for photo in photos:
                try:
                    file_extension = os.path.splitext(photo.filename)[1]
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    file_path = os.path.join(upload_dir, unique_filename)
                    
                    with open(file_path, "wb") as buffer:
                        shutil.copyfileobj(photo.file, buffer)
                    
                    relative_path = f"/uploads/photos/{repair_order.id}/{unique_filename}"
                    uploaded_photos.append(relative_path)
                    
                    attachment = Attachment(
                        file_name=photo.filename,
                        file_path=relative_path,
                        file_type="photo",
                        file_size=os.path.getsize(file_path),
                        mime_type=photo.content_type,
                        is_physical_damage=True,
                        repair_order_id=repair_order.id,
                        uploaded_by=current_user.id
                    )
                    db.add(attachment)
                except Exception as e:
                    logger.error(f"خطا در آپلود عکس {photo.filename}: {e}")
                    # ادامه می‌دهیم حتی اگر یک عکس مشکل داشته باشد
        
        # ============================================
        # ثبت تاریخچه وضعیت
        # ============================================
        
        history = StatusHistory(
            repair_order_id=repair_order.id,
            old_status=None,
            new_status=OrderStatus.REGISTERED,
            reason="ثبت پرونده جدید",
            operator_name=current_user.full_name,
            changed_by=current_user.id
        )
        db.add(history)
        
        # ============================================
        # Commit نهایی
        # ============================================
        
        db.commit()
        db.refresh(repair_order)
        
        logger.info(f"✅ پرونده نسخه ۲ ثبت شد: {repair_order.tracking_code}")
        
        return {
            "id": repair_order.id,
            "tracking_code": repair_order.tracking_code,
            "status": repair_order.status.value,
            "message": "پرونده با موفقیت ثبت شد",
            "photos_count": len(uploaded_photos),
            "boards_count": len(boards_data)
        }
    
    except HTTPException:
        # ✅ خطاهایی که خودمان ایجاد کردیم با پیام قابل فهم
        db.rollback()
        raise
    
    except json.JSONDecodeError:
        db.rollback()
        logger.error(f"خطا در pars کردن JSON: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail="اطلاعات ارسالی نامعتبر است. لطفاً دوباره تلاش کنید."
        )
    
    except Exception as e:
        db.rollback()
        logger.error(f"خطا در ثبت پرونده نسخه ۲: {e}")
        
        # ✅ تبدیل خطاهای رایج به پیام‌های قابل فهم
        error_msg = str(e)
        
        if "column" in error_msg and "does not exist" in error_msg:
            # خطای مربوط به ستون‌های دیتابیس
            raise HTTPException(
                status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="خطا در ساختار دیتابیس. لطفاً با پشتیبانی تماس بگیرید."
            )
        elif "not-null constraint" in error_msg:
            # خطای مربوط به فیلدهای اجباری
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="برخی از اطلاعات ضروری وارد نشده است. لطفاً فرم را کامل کنید."
            )
        elif "foreign key" in error_msg.lower():
            # خطای مربوط به روابط دیتابیس
            raise HTTPException(
                status_code=http_status.HTTP_400_BAD_REQUEST,
                detail="اطلاعات انتخاب شده معتبر نیست. لطفاً دوباره انتخاب کنید."
            )
        else:
            # خطای عمومی
            raise HTTPException(
                status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="خطا در ثبت پرونده. لطفاً دوباره تلاش کنید."
            )
# ============================================
# دریافت اطلاعات یک سایت با ID
# ============================================
@router.get("/sites/{site_id}")
def get_site(
    site_id: int,
    db: Session = Depends(get_db)
):
    """
    دریافت اطلاعات کامل یک سایت با ID
    """
    try:
        site = db.query(Site).filter(Site.id == site_id).first()
        if not site:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="سایت یافت نشد"
            )
        return site
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"خطا در دریافت سایت {site_id}: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="خطا در دریافت اطلاعات سایت"
        )
# ============================================
# دریافت برگه پذیرش - نسخه نهایی
# ============================================
# ============================================
# دریافت برگه پذیرش - بدون Jinja2
# ============================================
@router.get("/repair-orders/{order_id}/receipt")
async def get_receipt(
    order_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    دریافت برگه پذیرش پرونده - با HTML خالص
    """
    import json
    
    try:
        logger.info(f"🔍 دریافت برگه پذیرش برای پرونده {order_id}")
        
        # دریافت اطلاعات پرونده
        order = db.query(RepairOrder).options(
            joinedload(RepairOrder.customer),
            joinedload(RepairOrder.panel),
            joinedload(RepairOrder.site)
        ).filter(RepairOrder.id == order_id).first()
        
        if not order:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail="پرونده یافت نشد"
            )
        
        # دریافت بردها
        boards = db.query(Board).filter(Board.repair_order_id == order_id).all()
        
        # آماده‌سازی داده‌ها
        data = {
            "order_id": order.id,
            "tracking_code": order.tracking_code,
            "status": order.status.value if hasattr(order.status, 'value') else str(order.status),
            "reception_date": order.reception_date.strftime("%Y-%m-%d %H:%M") if order.reception_date else "",
            "operator": order.operator_name or "نامشخص",
            "customer_complaint": order.customer_complaint or "",
            "physical_damages": order.physical_damages if order.physical_damages else [],
            "physical_description": order.physical_description or "",
            "accessories": order.accessories if order.accessories else [],
            "accessories_description": order.accessories_description or "",
            "customer": {
                "name": order.customer.name if order.customer else "",
                "phone": order.customer.phone if order.customer else "",
                "address": order.customer.address if order.customer else "",
                "company": order.customer.company if order.customer else ""
            },
            "panel": {
                "brand": order.panel.brand if order.panel else "",
                "model": order.panel.model if order.panel else "",
                "serial_number": order.panel.serial_number if order.panel else "",
                "part_number": order.panel.part_number if order.panel else ""
            },
            "site": {
                "name": order.site.name if order.site else "",
                "address": order.site.address if order.site else "",
                "type": order.site.type if order.site else ""
            },
            "boards": [
                {
                    "board_type": b.board_type.value if hasattr(b.board_type, 'value') else str(b.board_type),
                    "part_number": b.part_number or "",
                    "serial_number": b.serial_number or "",
                    "revision": b.revision or ""
                }
                for b in boards
            ]
        }
        
        # ============================================
        # ساخت HTML خالص
        # ============================================
        html_content = f"""
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>برگه پذیرش - {data['tracking_code']}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.rtl.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        @media print {{
            .no-print {{ display: none !important; }}
            body {{ background: white !important; padding: 20px; }}
            .receipt-card {{ box-shadow: none !important; border: 1px solid #ddd !important; }}
        }}
        body {{ font-family: 'Tahoma', sans-serif; background: #f8f9fa; padding: 20px; }}
        .receipt-card {{ max-width: 800px; margin: 0 auto; background: white; border-radius: 15px; box-shadow: 0 5px 30px rgba(0,0,0,0.1); padding: 30px; }}
        .receipt-header {{ text-align: center; border-bottom: 2px solid #0d6efd; padding-bottom: 20px; margin-bottom: 20px; }}
        .receipt-header h2 {{ color: #0d6efd; }}
        .tracking-code {{ font-size: 24px; font-weight: bold; color: #198754; background: #e7f5e9; padding: 5px 20px; border-radius: 25px; display: inline-block; margin-top: 10px; }}
        .info-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px dashed #eee; }}
        .info-row .label {{ font-weight: bold; color: #555; min-width: 120px; }}
        .info-row .value {{ color: #333; text-align: left; }}
        .section-title {{ color: #0d6efd; font-weight: bold; margin-top: 25px; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #0d6efd; }}
        .damage-badge {{ display: inline-block; padding: 5px 15px; border-radius: 20px; margin: 3px; background: #f8d7da; color: #721c24; }}
        .accessory-badge {{ display: inline-block; padding: 5px 15px; border-radius: 20px; margin: 3px; background: #d1ecf1; color: #0c5460; }}
        .board-item {{ background: #f8f9fa; border-radius: 10px; padding: 10px 15px; margin-bottom: 10px; border: 1px solid #e9ecef; }}
        .signature-box {{ border: 1px solid #ddd; border-radius: 10px; padding: 20px; margin-top: 30px; text-align: center; }}
        .signature-box .sig-line {{ width: 200px; border-bottom: 2px solid #333; margin: 20px auto 5px; }}
        .print-btn {{ position: fixed; bottom: 20px; left: 50%; transform: translateX(-50%); z-index: 1000; padding: 12px 40px; font-size: 18px; border-radius: 30px; box-shadow: 0 4px 15px rgba(0,0,0,0.2); }}
        .back-btn {{ position: fixed; bottom: 80px; left: 50%; transform: translateX(-50%); z-index: 1000; padding: 10px 30px; font-size: 16px; border-radius: 30px; width: 200px; }}
        .empty-value {{ color: #adb5bd; font-style: italic; }}
    </style>
</head>
<body>
    <div class="receipt-card">
        <!-- هدر -->
        <div class="receipt-header">
            <h2><i class="fas fa-fire-extinguisher"></i> سیستم مدیریت تعمیرات</h2>
            <h4>برگه پذیرش پرونده</h4>
            <div class="tracking-code"><i class="fas fa-barcode"></i> {data['tracking_code']}</div>
            <div class="mt-2">
                <span class="badge bg-secondary">تاریخ: {data['reception_date']}</span>
                <span class="badge bg-info ms-2">وضعیت: {data['status']}</span>
                <span class="badge bg-primary ms-2">اپراتور: {data['operator']}</span>
            </div>
        </div>

        <!-- مشتری -->
        <h5 class="section-title"><i class="fas fa-user"></i> اطلاعات مشتری</h5>
        <div class="info-row"><span class="label">نام:</span><span class="value">{data['customer']['name'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">شماره تماس:</span><span class="value">{data['customer']['phone'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">آدرس:</span><span class="value">{data['customer']['address'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">شرکت:</span><span class="value">{data['customer']['company'] or 'ثبت نشده'}</span></div>

        <!-- محل نصب -->
        <h5 class="section-title"><i class="fas fa-map-marker-alt"></i> محل نصب</h5>
        <div class="info-row"><span class="label">نام محل:</span><span class="value">{data['site']['name'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">نوع:</span><span class="value">{data['site']['type'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">آدرس:</span><span class="value">{data['site']['address'] or 'ثبت نشده'}</span></div>

        <!-- پنل -->
        <h5 class="section-title"><i class="fas fa-microchip"></i> اطلاعات پنل</h5>
        <div class="info-row"><span class="label">برند:</span><span class="value">{data['panel']['brand'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">مدل:</span><span class="value">{data['panel']['model'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">سریال نامبر:</span><span class="value">{data['panel']['serial_number'] or 'ثبت نشده'}</span></div>
        <div class="info-row"><span class="label">پارت نامبر:</span><span class="value">{data['panel']['part_number'] or 'ثبت نشده'}</span></div>

        <!-- بردها -->
        <h5 class="section-title"><i class="fas fa-layer-group"></i> ساختار بردها</h5>
        {''.join([f'<div class="board-item"><strong>{b["board_type"]}</strong> <span class="badge bg-secondary">PN: {b["part_number"]}</span> <span class="badge bg-info">SN: {b["serial_number"]}</span></div>' for b in data['boards']]) or '<p class="text-muted">هیچ بردی ثبت نشده است</p>'}

        <!-- وضعیت ظاهری -->
        <h5 class="section-title"><i class="fas fa-eye"></i> وضعیت ظاهری</h5>
        <div>
            {''.join([f'<span class="damage-badge">{d}</span>' for d in data['physical_damages']]) or '<span class="text-muted">ثبت نشده</span>'}
        </div>
        {f'<div class="mt-2"><strong>توضیحات:</strong> {data["physical_description"]}</div>' if data['physical_description'] else ''}

        <!-- متعلقات -->
        <h5 class="section-title"><i class="fas fa-box"></i> متعلقات</h5>
        <div>
            {''.join([f'<span class="accessory-badge">{a}</span>' for a in data['accessories']]) or '<span class="text-muted">ثبت نشده</span>'}
        </div>
        {f'<div class="mt-2"><strong>توضیحات:</strong> {data["accessories_description"]}</div>' if data['accessories_description'] else ''}

        <!-- شرح مشتری -->
        <h5 class="section-title"><i class="fas fa-pen"></i> شرح مشتری</h5>
        <div class="alert alert-light border">{data['customer_complaint'] or 'ثبت نشده'}</div>

        <!-- امضاها -->
        <div class="signature-box">
            <div class="row">
                <div class="col-6"><div class="sig-line"></div><small>امضای مشتری</small></div>
                <div class="col-6"><div class="sig-line"></div><small>امضای پذیرش</small></div>
            </div>
            <div class="mt-3"><small class="text-muted">تاریخ چاپ: {data['reception_date']}</small></div>
        </div>

        <!-- فوتر -->
        <div class="text-center mt-4 text-muted" style="font-size: 12px;">
            <i class="fas fa-fire-extinguisher"></i> سیستم مدیریت تعمیرات تجهیزات اعلام و اطفا حریق
        </div>
    </div>

    <button class="btn btn-primary btn-lg print-btn no-print" onclick="window.print()">
        <i class="fas fa-print"></i> چاپ برگه
    </button>
    <a href="/order/{data['order_id']}" class="btn btn-secondary back-btn no-print">
        <i class="fas fa-arrow-right"></i> بازگشت
    </a>

    <script>
        document.addEventListener('keydown', function(e) {{
            if ((e.ctrlKey || e.metaKey) && e.key === 'p') {{
                e.preventDefault();
                window.print();
            }}
        }});
    </script>
</body>
</html>
        """
        
        # ============================================
        # برگرداندن HTML
        # ============================================
        return HTMLResponse(content=html_content)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ خطا در دریافت برگه پذیرش: {e}")
        raise HTTPException(
            status_code=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"خطا در دریافت اطلاعات برگه: {str(e)}"
        )